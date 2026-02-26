"""Leads API endpoints."""
import csv
import io
from typing import List, Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy import select, func, delete
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

import structlog

from app.dependencies import get_db
from app.models.lead import Lead, LeadIntelligence
from app.models.draft import Draft
from app.models.in_sequence import CampaignLead
from app.models.event import EmailEvent
from app.schemas.lead import (
    LeadCreate,
    LeadUpdate,
    LeadResponse,
    LeadDetailResponse,
    EmailEventResponse,
    LeadListResponse,
    BulkLeadImport,
    BulkImportResult,
)
from fastapi import BackgroundTasks

logger = structlog.get_logger()
router = APIRouter()


@router.post("", response_model=LeadResponse, status_code=201)
async def create_lead(
    lead_in: LeadCreate,
    db: AsyncSession = Depends(get_db),
):
    """Create a new lead."""
    # Check for duplicate
    if lead_in.external_id:
        stmt = select(Lead).where(Lead.external_id == lead_in.external_id)
        existing = await db.execute(stmt)
        if existing.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="Lead with this external_id already exists")
    
    lead = Lead(**lead_in.model_dump())
    db.add(lead)
    await db.flush()
    await db.refresh(lead)
    
    print(f"\n[DEBUG] >>> LEAD CREATED: {lead.company_name} ({lead.email})")
    return lead


@router.get("/{lead_id}", response_model=LeadDetailResponse)
async def get_lead(
    lead_id: UUID,
    db: AsyncSession = Depends(get_db),
):
    """Get detailed lead information including events."""
    stmt = (
        select(Lead)
        .options(selectinload(Lead.events))
        .options(selectinload(Lead.intelligence))
        .where(Lead.id == lead_id)
    )
    result = await db.execute(stmt)
    lead = result.scalar_one_or_none()
    
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    return lead


@router.post("/bulk", response_model=BulkImportResult)
async def bulk_import_leads(
    import_data: BulkLeadImport,
    db: AsyncSession = Depends(get_db),
):
    """Bulk import leads from list."""
    created = 0
    skipped = 0
    errors = []
    
    for i, lead_data in enumerate(import_data.leads):
        try:
            # Check for duplicate
            if lead_data.external_id:
                stmt = select(Lead).where(Lead.external_id == lead_data.external_id)
                existing = await db.execute(stmt)
                if existing.scalar_one_or_none():
                    skipped += 1
                    continue
            
            lead = Lead(**lead_data.model_dump())
            db.add(lead)
            created += 1
            
        except Exception as e:
            errors.append({"index": i, "error": str(e)})
    
    await db.flush()
    
    return BulkImportResult(created=created, skipped=skipped, errors=errors)


# Column name mappings (lowercase keys for case-insensitive matching)
COLUMN_MAPPINGS = {
    "company_name": "company_name",
    "company name": "company_name",
    "company": "company_name",
    "company_domain": "company_domain",
    "domain": "company_domain",
    "website": "company_domain",
    "email": "email",
    "email address": "email",
    "first_name": "first_name",
    "first name": "first_name",
    "firstname": "first_name",
    "last_name": "last_name",
    "last name": "last_name",
    "lastname": "last_name",
    "linkedin_url": "linkedin_url",
    "linkedin": "linkedin_url",
    "linkedin url": "linkedin_url",
    "industry": "industry",
    "mobile": "mobile",
    "phone": "mobile",
    "phone number": "mobile",
    "region": "region",
    "persona": "persona",
    "external_id": "external_id",
}


def _normalize_column_name(col: str) -> Optional[str]:
    """Map CSV/Excel column name to lead field name."""
    return COLUMN_MAPPINGS.get(col.lower().strip())


def _parse_csv(content: str) -> List[dict]:
    """Parse CSV content into list of row dicts."""
    reader = csv.DictReader(io.StringIO(content))
    rows = []
    for row in reader:
        mapped = {}
        for col, value in row.items():
            if value and value.strip():
                field = _normalize_column_name(col)
                if field:
                    mapped[field] = value.strip()
        if mapped.get("company_name") and mapped.get("company_domain"):
            rows.append(mapped)
    return rows


def _parse_excel(file_content: bytes) -> List[dict]:
    """Parse Excel file into list of row dicts."""
    from openpyxl import load_workbook
    
    wb = load_workbook(filename=io.BytesIO(file_content), read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    
    # Get headers from first row
    headers = next(rows_iter, None)
    if not headers:
        return []
    
    # Map headers to field names
    header_mapping = []
    for h in headers:
        if h:
            field = _normalize_column_name(str(h))
            header_mapping.append(field)
        else:
            header_mapping.append(None)
    
    rows = []
    for row in rows_iter:
        mapped = {}
        for i, value in enumerate(row):
            if i < len(header_mapping) and header_mapping[i] and value:
                mapped[header_mapping[i]] = str(value).strip()
        if mapped.get("company_name") and mapped.get("company_domain"):
            rows.append(mapped)
    
    return rows


@router.post("/upload", response_model=BulkImportResult)
async def upload_leads_file(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """Upload CSV or Excel file to bulk import leads."""
    # Validate file type
    filename = file.filename or ""
    if not (filename.endswith(".csv") or filename.endswith(".xlsx")):
        raise HTTPException(
            status_code=400,
            detail="Invalid file type. Please upload a CSV (.csv) or Excel (.xlsx) file."
        )
    
    # Read file content
    content = await file.read()
    
    # Parse based on file type
    try:
        if filename.endswith(".csv"):
            rows = _parse_csv(content.decode("utf-8"))
        else:
            rows = _parse_excel(content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")
    
    if not rows:
        raise HTTPException(
            status_code=400,
            detail="No valid leads found. Ensure file has 'company_name' and 'company_domain' columns."
        )
    
    # Import leads
    created = 0
    skipped = 0
    errors = []
    
    for i, row_data in enumerate(rows):
        try:
            # Check for duplicate by email
            if row_data.get("email"):
                stmt = select(Lead).where(Lead.email == row_data["email"])
                existing = await db.execute(stmt)
                if existing.scalar_one_or_none():
                    skipped += 1
                    continue
            
            # Set defaults
            row_data.setdefault("personalization_mode", "medium")
            
            lead = Lead(**row_data)
            db.add(lead)
            created += 1
            
        except Exception as e:
            errors.append({"row": i + 2, "error": str(e)})  # +2 for 1-indexed + header
    
    await db.flush()
    
    return BulkImportResult(created=created, skipped=skipped, errors=errors)


@router.get("", response_model=LeadListResponse)
async def list_leads(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=100),
    status: Optional[str] = None,
    industry: Optional[str] = None,
    min_score: Optional[float] = Query(None, ge=0, le=1),
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    """List leads with pagination and filtering."""
    # Base query
    stmt = select(Lead)
    count_stmt = select(func.count(Lead.id))
    
    # Apply filters
    if status:
        if status == "inprogress":
            # Outreach phase
            inprogress_statuses = ["sequencing", "started", "inprogress", "contacted", "mail_sent"]
            stmt = stmt.where(Lead.status.in_(inprogress_statuses))
            count_stmt = count_stmt.where(Lead.status.in_(inprogress_statuses))
        elif status == "qualified":
            # Qualified but outreach not yet started
            stmt = stmt.where(Lead.status == "qualified")
            count_stmt = count_stmt.where(Lead.status == "qualified")
        elif status == "completed":
            # Finished phase
            completed_statuses = ["replied", "converted", "disqualified", "completed"]
            stmt = stmt.where(Lead.status.in_(completed_statuses))
            count_stmt = count_stmt.where(Lead.status.in_(completed_statuses))
        elif status == "all_contacted":
            # Everything that has reached the contact stage
            contacted_statuses = ["contacted", "mail_sent", "replied", "converted", "completed"]
            stmt = stmt.where(Lead.status.in_(contacted_statuses))
            count_stmt = count_stmt.where(Lead.status.in_(contacted_statuses))
        elif status == "new":
            # New leads, those never started, or those currently being researched
            stmt = stmt.where(Lead.status.in_(["new", "not_started", "researching"]))
            count_stmt = count_stmt.where(Lead.status.in_(["new", "not_started", "researching"]))
        else:
            # Exact match for researching, qualified, etc.
            stmt = stmt.where(Lead.status == status)
            count_stmt = count_stmt.where(Lead.status == status)
    
    if industry:
        stmt = stmt.where(Lead.industry.ilike(f"%{industry}%"))
        count_stmt = count_stmt.where(Lead.industry.ilike(f"%{industry}%"))
    
    if min_score is not None:
        stmt = stmt.where(Lead.composite_score >= min_score)
        count_stmt = count_stmt.where(Lead.composite_score >= min_score)
    
    if search:
        search_filter = (
            Lead.company_name.ilike(f"%{search}%") |
            Lead.email.ilike(f"%{search}%") |
            Lead.first_name.ilike(f"%{search}%") |
            Lead.last_name.ilike(f"%{search}%")
        )
        stmt = stmt.where(search_filter)
        count_stmt = count_stmt.where(search_filter)
    
    # Get total count
    total_result = await db.execute(count_stmt)
    total = total_result.scalar() or 0
    
    # Apply pagination
    offset = (page - 1) * page_size
    stmt = stmt.offset(offset).limit(page_size)
    stmt = stmt.order_by(Lead.composite_score.desc().nullslast(), Lead.created_at.desc())
    
    # Efficiently check has_replied in a single subquery or post-fetch
    result = await db.execute(stmt)
    leads = result.scalars().all()
    
    if leads:
        lead_ids = [l.id for l in leads]
        replied_stmt = select(EmailEvent.lead_id).where(
            (EmailEvent.lead_id.in_(lead_ids)) & (EmailEvent.event_type == "replied")
        ).group_by(EmailEvent.lead_id)
        replied_result = await db.execute(replied_stmt)
        replied_ids = set(replied_result.scalars().all())
        
        for lead in leads:
            lead.has_replied = lead.id in replied_ids
    
    return LeadListResponse(
        items=leads,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size if page_size > 0 else 0,
    )


@router.get("/{lead_id}", response_model=LeadDetailResponse)
async def get_lead(
    lead_id: UUID,
    db: AsyncSession = Depends(get_db),
):
    """Get lead with intelligence."""
    stmt = (
        select(Lead)
        .options(selectinload(Lead.intelligence))
        .options(selectinload(Lead.events))
        .where(Lead.id == lead_id)
    )
    result = await db.execute(stmt)
    lead = result.scalar_one_or_none()
    
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    # Add has_replied flag by checking EmailEvents
    reply_stmt = select(func.count(EmailEvent.id)).where(
        (EmailEvent.lead_id == lead.id) & (EmailEvent.event_type == "replied")
    )
    reply_result = await db.execute(reply_stmt)
    lead.has_replied = reply_result.scalar() > 0
    
    return lead


@router.get("/{lead_id}/intelligence", response_model=dict)
async def get_lead_intelligence(
    lead_id: UUID,
    db: AsyncSession = Depends(get_db),
):
    """Get detailed research/intelligence data for a lead."""
    from app.models.lead import LeadIntelligence
    
    # Get lead with basic info
    lead_stmt = select(Lead).where(Lead.id == lead_id)
    lead_result = await db.execute(lead_stmt)
    lead = lead_result.scalar_one_or_none()
    
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    # Get intelligence
    intel_stmt = select(LeadIntelligence).where(LeadIntelligence.lead_id == lead_id)
    intel_result = await db.execute(intel_stmt)
    intelligence = intel_result.scalar_one_or_none()
    
    if not intelligence:
        return {
            "lead": {
                "id": str(lead.id),
                "company_name": lead.company_name,
                "company_domain": lead.company_domain,
                "first_name": lead.first_name,
                "last_name": lead.last_name,
                "email": lead.email,
                "status": lead.status,
            },
            "researched": False,
            "message": "No research has been conducted for this lead yet."
        }
    
    return {
        "lead": {
            "id": str(lead.id),
            "company_name": lead.company_name,
            "company_domain": lead.company_domain,
            "first_name": lead.first_name,
            "last_name": lead.last_name,
            "email": lead.email,
            "status": lead.status,
        },
        "researched": True,
        "researched_at": intelligence.researched_at.isoformat() if intelligence.researched_at else None,
        "is_stale": intelligence.is_stale,
        "scores": {
            "fit_score": float(lead.fit_score) if lead.fit_score else None,
            "readiness_score": float(lead.readiness_score) if lead.readiness_score else None,
            "intent_score": float(lead.intent_score) if lead.intent_score else None,
            "composite_score": float(lead.composite_score) if lead.composite_score else None,
        },
        "your_company": {
            "services": intelligence.your_services,
            "proof_points": intelligence.your_proof_points,
            "positioning": intelligence.your_positioning,
        },
        "lead_analysis": {
            "offerings": intelligence.lead_offerings,
            "pain_indicators": intelligence.lead_pain_indicators,
            "buying_signals": intelligence.lead_buying_signals,
            "tech_stack": intelligence.lead_tech_stack,
        },
        "linkedin": {
            "role": intelligence.linkedin_role,
            "seniority": intelligence.linkedin_seniority,
            "topics_30d": intelligence.linkedin_topics_30d,
            "job_change_days": intelligence.linkedin_job_change_days,
            "likely_initiatives": intelligence.linkedin_likely_initiatives,
            # Enhanced LinkedIn intelligence
            "decision_power": intelligence.linkedin_decision_power,
            "budget_authority": intelligence.linkedin_budget_authority,
            "lead_score": intelligence.linkedin_lead_score,
            "cold_email_hooks": intelligence.cold_email_hooks or [],
            "opening_line": intelligence.opening_line,
        },
        "triggers": intelligence.triggers or [],
        "insights": {
            "pain_hypotheses": intelligence.pain_hypotheses,
            "best_angle": intelligence.best_angle,
        },
        "events": [
            {
                "id": str(event.id),
                "event_type": event.event_type,
                "touch_number": event.touch_number,
                "created_at": event.created_at.isoformat(),
                "title": event.title or (event.draft.selected_subject if event.draft else None),
                "body": event.body or (event.draft.body if event.draft else None),
                "reply_sentiment": event.reply_sentiment,
                "reply_intent": event.reply_intent,
            }
            for event in (await db.scalars(
                select(EmailEvent)
                .options(selectinload(EmailEvent.draft))
                .where(EmailEvent.lead_id == lead_id)
                .order_by(EmailEvent.created_at.desc())
            )).all()
        ],
    }


@router.patch("/{lead_id}", response_model=LeadResponse)
async def update_lead(
    lead_id: UUID,
    lead_in: LeadUpdate,
    db: AsyncSession = Depends(get_db),
):
    """Update a lead."""
    stmt = select(Lead).where(Lead.id == lead_id)
    result = await db.execute(stmt)
    lead = result.scalar_one_or_none()
    
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    # Update fields
    update_data = lead_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(lead, field, value)
    
    await db.flush()
    await db.refresh(lead)
    
    return lead


@router.delete("/{lead_id}", status_code=204)
async def delete_lead(
    lead_id: UUID,
    db: AsyncSession = Depends(get_db),
):
    """Delete a lead (GDPR compliance)."""
    stmt = select(Lead).where(Lead.id == lead_id)
    result = await db.execute(stmt)
    lead = result.scalar_one_or_none()
    
    
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    # Explicitly delete related records using bulk delete to avoid ORM/FK constraint issues
    await db.execute(delete(LeadIntelligence).where(LeadIntelligence.lead_id == lead_id))
    await db.execute(delete(Draft).where(Draft.lead_id == lead_id))
    await db.execute(delete(CampaignLead).where(CampaignLead.lead_id == lead_id))
    await db.execute(delete(EmailEvent).where(EmailEvent.lead_id == lead_id))
    
    await db.delete(lead)
    await db.commit()


@router.post("/{lead_id}/research", status_code=202)
async def trigger_research(
    lead_id: UUID,
    background_tasks: BackgroundTasks,
    force_refresh: bool = False,
    db: AsyncSession = Depends(get_db),
):
    """Trigger research job for a lead."""
    stmt = select(Lead).where(Lead.id == lead_id)
    result = await db.execute(stmt)
    lead = result.scalar_one_or_none()
    
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    # Check if already researched recently — but only skip if research was successful (non-zero score)
    from datetime import datetime, timedelta
    if lead.researched_at and not force_refresh:
        has_scores = lead.composite_score is not None and float(lead.composite_score) > 0
        if has_scores and lead.researched_at > datetime.utcnow() - timedelta(days=30):
            return {
                "status": "skipped",
                "reason": "Recently researched with valid scores",
                "researched_at": lead.researched_at.isoformat(),
                "composite_score": float(lead.composite_score),
            }
    
    # Update status to researching immediately
    try:
        old_status = lead.status
        lead.status = "researching"
        # We use commit() here to ensure the state is saved before the background task or frontend poll hits
        await db.commit()
        await db.refresh(lead)
        
        logger.info("Lead status updated to researching", lead_id=str(lead_id), previous_status=old_status)
        
        # Trigger Background Task for async research (Bypassing Celery to avoid Redis connection limits)
        from app.services.research import run_research_background
        background_tasks.add_task(run_research_background, str(lead_id))
        logger.info("Research background task queued", lead_id=str(lead_id))
        
        return {
            "status": "queued",
            "lead_id": str(lead_id),
            "job_id": str(lead_id),
            "current_status": lead.status,
        }
    except Exception as e:
        logger.error("Failed to trigger research", error=str(e), lead_id=str(lead_id))
        # Rollback if the status update failed
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error while starting research: {str(e)}")
